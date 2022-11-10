import numpy
import openpyxl

# Set up




Average = 50
Variance = 5
N = [10,30,50,70,90,100,130,160,180,240,260,300,400,500,700,900,1000]
numberOfLists = 2000

workbook = openpyxl.load_workbook("Data.xlsx")
cutOffResults = workbook.worksheets[0]
adaptiveResults = workbook.worksheets[1]

cutOffResults.cell(1,1,"N")
cutOffResults.cell(1,2,"Optimal CutOff")
cutOffResults.cell(1,3,"Optimal Benchmark")
cutOffResults.cell(1,4,"Expectation")

adaptiveResults.cell(1,1,"N")
adaptiveResults.cell(1,2,"Optimal Parameter")
adaptiveResults.cell(1,3,"Optimal Benchmark")
adaptiveResults.cell(1,4,"Expectation")

worksheetRow = 2
worksheetRow_adaptive = 2

# loop each candidates group
for numberOfCandidates in N:
    print(numberOfCandidates)
    cutOffResults.cell(worksheetRow, 1, numberOfCandidates)
    adaptiveResults.cell(worksheetRow_adaptive,1,numberOfCandidates)
    cutOffExpectation = []
    adaptiveExpectation = []
    cutOffBenchmarkPool = [[] for _ in range(numberOfCandidates - 1)]
    # adaptiveBenchmarkPool = [[] for _ in range(501)]
    cutOffPool = [[] for _ in range(numberOfCandidates - 1)]
    adaptivePool = [[] for _ in range(501)]
    parameter = 0

    # loop each cut off from 1 to numberOfCandidates - 1
    for cutOff in range(1,numberOfCandidates):

        # Generate Normal Distribution
        for i in range(numberOfLists):
            normalList = numpy.round(numpy.random.normal(Average, Variance, size=(numberOfCandidates,1)),3)

            # benchmark = the best candidate among normalList[0] ~ normalList[cutOff-1]
            # normalList[0] ~ normalList[cutOff-1] includes normalList[0] and normalList[cutOff-1]
            benchmark = max(normalList[:cutOff])

            # add this benchmark current cutOffBenchmarkPool
            cutOffBenchmarkPool[cutOff - 1].append(benchmark)

            # candidateSelected = after cutOff, the first candidate who is better than benchmark
            for candidateIndex in range(cutOff,numberOfCandidates):
                if normalList[candidateIndex] > benchmark:
                    candidateSelected = normalList[candidateIndex]
                    break
                # After cutOff, when it comes to the last candidate.
                # if previously no one was better than benchmark, select the last candidate
                if candidateIndex == numberOfCandidates - 1:
                    candidateSelected = normalList[-1]

            # Add this candidate to current cutOff pool
            cutOffPool[cutOff - 1].append(candidateSelected)

    #Loop parameter
    while parameter <= 5:

        # Generate Normal Distribution
        for i in range(numberOfLists):
            normalList = numpy.round(numpy.random.normal(Average, Variance, size=(numberOfCandidates, 1)), 3)
            
            adaptiveBenchmark = Average + parameter * Variance
            
            for candidateIndex in range(numberOfCandidates):
                if normalList[candidateIndex] > adaptiveBenchmark:
                    candidateSelected = normalList[candidateIndex]
                    break
                if candidateIndex == numberOfCandidates - 1:
                    candidateSelected = normalList[-1]

            adaptivePool[int(parameter * 100)].append(candidateSelected)

        parameter += 0.01

    #Calculate the expectation value for each cutOff
    for i in cutOffPool:
        cutOffExpectation.append(numpy.mean(i))

    # Calculate the expectation value for each parameter
    for i in adaptivePool:
        adaptiveExpectation.append(numpy.mean(i))

    # Get the max expectation
    maxExpectation = max(cutOffExpectation)
    maxAdaptiveExpectation = max(adaptiveExpectation)

    # save all the cutOff with the best expectation
    optimalCutOff = []
    for i in range(numberOfCandidates - 1):
        if cutOffExpectation[i] == maxExpectation:
            optimalCutOff.append(i + 1)

    # save all the parameter with the best expectation
    optimalParameter = []
    for i in range(501):
        if adaptiveExpectation[i] == maxAdaptiveExpectation:
            optimalParameter.append(i/100)

    for i in optimalCutOff:
        cutOffResults.cell(worksheetRow, 2, i)
        cutOffResults.cell(worksheetRow, 3, numpy.mean(cutOffBenchmarkPool[i - 1]))
        cutOffResults.cell(worksheetRow, 4, maxExpectation)
        worksheetRow += 1

    for i in optimalParameter:
        adaptiveResults.cell(worksheetRow_adaptive, 2, i)
        adaptiveResults.cell(worksheetRow_adaptive, 3, Average + i * Variance)
        adaptiveResults.cell(worksheetRow_adaptive, 4, maxAdaptiveExpectation)
        worksheetRow_adaptive += 1

workbook.save(filename='Data.xlsx')



